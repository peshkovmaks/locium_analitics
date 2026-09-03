#!/usr/bin/env python3
"""Inspect downloaded Yandex Market XLSX reports and print structural summary."""

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

FILES = {
    "united_marketplace_services": Path(__file__).resolve().parent / "ym_united_marketplace_services.xlsx",
    "united_orders": Path(__file__).resolve().parent / "ym_united_orders.xlsx",
}


def find_header_row(ws):
    """Return the first row index that looks like a header (has 'ID бизнес-аккаунта' or 'Ваш SKU')."""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("id бизнес-аккаунта" in c for c in cells) or any("ваш sku" in c for c in cells):
            return i
    return 1


def summarize_sheet(ws, max_data_rows=5):
    header_row = find_header_row(ws)
    headers = []
    if header_row <= ws.max_row:
        headers = [cell.value for cell in ws[header_row]]
    data_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if len(data_rows) >= max_data_rows:
            break
        if any(v is not None and str(v).strip() for v in row):
            data_rows.append(row)
    return {
        "header_row": header_row,
        "headers": headers,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "sample_rows": data_rows,
    }


def main():
    summary = {}
    for name, path in FILES.items():
        if not path.exists():
            print(f"Missing {path}")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets[sheet_name] = summarize_sheet(ws)
        summary[name] = sheets
        wb.close()

    out_path = Path(__file__).resolve().parent / "ym_xlsx_inspection.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)

    print(f"Inspection written to {out_path}")
    # Console summary
    for report_name, sheets in summary.items():
        print(f"\n=== {report_name} ===")
        for sheet_name, info in sheets.items():
            print(f"\n  Sheet: {sheet_name} ({info['max_row']} rows x {info['max_column']} cols)")
            print(f"    Header row: {info['header_row']}")
            headers = info["headers"]
            print(f"    Headers ({len([h for h in headers if h])} non-empty):")
            for idx, h in enumerate(headers, start=1):
                if h:
                    print(f"      {idx}: {h}")
            print(f"    Sample data rows: {len(info['sample_rows'])}")
            for row in info["sample_rows"][:2]:
                pairs = [(headers[i], v) for i, v in enumerate(row) if i < len(headers) and headers[i] and v is not None]
                print(f"      {dict(pairs)}")


if __name__ == "__main__":
    main()
